import json
import os
from openai import Client
import time
from GuidelineService import GuidelineService
from openpyxl import load_workbook

def extract_json(text):
    # Find the start of the JSON block
    start_index = text.find("{")
    if start_index == -1:
        return "JSON block not found."

    # Find the end of the JSON block
    end_index = text.rfind("}")
    if end_index == -1:
        return "JSON block not found."

    # Extract and return the JSON block
    return text[start_index:end_index + 1]

def get_assessment(guideline_title, reco, subsections):
    print("Processing recommendation %s" % reco['number'])
    reco_type = reco['type_of_recommendation']['id']
    reco_number = reco['number'].replace(".", "_")
    if reco_number in cb_reco_list or 'statement' in reco_type or 'evidence' in reco_type:
        return (None, None)
    reco_text = html2text.handle(reco['text']).replace("\r\n", " ").replace("\r", " ").replace("\n", " ")
    reco_type = "consensus-based"
    reco_loe = "N/A"
    reco_literature = "none"
    if 'statement' in reco['type_of_recommendation']['id']:
        reco_gor = "N/A"
    else:
        reco_gor = parse_gor(reco_text)

    param_list = ["clarityAndActionability",
                  "necessityInHealthcare",
                  "netPositiveConsequence",
                  "opportunityCost",
                  "rationaleForIndirectEvidence",
                  "justificationForStrengthOfRecommendation"
                  ]

    messag_content = "GoR: %s, Rtext: %s, Btext: %s" % (reco_gor, reco_text, reco_background_text)
    thread = client.beta.threads.create(
        messages=[
            {
                "role": "user",
                "content": messag_content,
            }
        ]
    )

    run = client.beta.threads.runs.create(
        thread_id=thread.id,
        assistant_id="asst_u90HBHegSrIAmyfBMo2j6sbN",
    )

    while run.status != "completed":
        run = client.beta.threads.runs.retrieve(
            thread_id=thread.id,
            run_id=run.id,
            timeout=10
        )

        time.sleep(2)

    messages = client.beta.threads.messages.list(
        thread_id=thread.id
    )

    text_val = extract_json(messages.data[0].content[0].text.value)
    etd_json = json.loads(text_val)

    result = "%s|%s|%s|%s|%s|%s|%s|" % (guideline_title, reco_number, reco_type, reco_gor, reco_loe, reco_text, reco_background_text)
    for param in param_list:
        result += "%s|%s|" % (etd_json['etdAssessment'][param]['explanation'], etd_json['etdAssessment'][param]['grade'])

    return (reco_type, result)

def assess_cb(reco_text, background_text, reco_gor):
    messag_content = "GoR: %s, Rtext: %s, Btext: %s" % (reco_gor, reco_text, background_text)
    thread = client.beta.threads.create(
        messages=[
            {
                "role": "user",
                "content": messag_content,
            }
        ]
    )

    run = client.beta.threads.runs.create(
        thread_id=thread.id,
        assistant_id="asst_u90HBHegSrIAmyfBMo2j6sbN",
    )

    while run.status != "completed":
        run = client.beta.threads.runs.retrieve(
            thread_id=thread.id,
            run_id=run.id,
            timeout=10
        )

        time.sleep(2)

    messages = client.beta.threads.messages.list(
        thread_id=thread.id
    )

    text_val = extract_json(messages.data[0].content[0].text.value)
    etd_json = json.loads(text_val)

    return etd_json['etdAssessment']

def assess_eb(reco_text, background_text, reco_gor):
    messag_content = "GoR: %s, Rtext: %s, Btext: %s" % (reco_gor, reco_text, background_text)
    thread = client.beta.threads.create(
        messages=[
            {
                "role": "user",
                "content": messag_content,
            }
        ]
    )

    run = client.beta.threads.runs.create(
        thread_id=thread.id,
        assistant_id="asst_nLffgXTlb9y6lPUyR778pCc0",
    )

    while run.status != "completed":
        run = client.beta.threads.runs.retrieve(
            thread_id=thread.id,
            run_id=run.id,
            timeout=10
        )

        time.sleep(2)

    messages = client.beta.threads.messages.list(
        thread_id=thread.id
    )

    text_val = extract_json(messages.data[0].content[0].text.value)
    etd_json = json.loads(text_val)

    return etd_json['etdAssessment']

client = Client(api_key="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")
service = GuidelineService()

source_dir = "//dkg-dc-01/Daten/Daten/Arbeitsverzeichnisse/OL/Kongresse_Vorträge_Publikationen/Guidelines International Network (G-I-N)/G-I-N 2024/EtD-Prüfung durch KI"
source_file = "2024-02-13_Bewertungsbogen KI_Template.xlsx"
source_path = os.path.join(source_dir, source_file)
target_file = "2024-02-19_Bewertungsbogen KI 1.xlsx"
target_path = os.path.join(source_dir, target_file)

workbook = load_workbook(source_path)
for sheet_name in workbook.sheetnames:
    sheet = workbook[sheet_name]
    for row in sheet.iter_rows(min_row=2, values_only=False):
        print("Row %s..." % row[0])
        reco_text = row[7].value
        background_text = row[8].value
        reco_gor = row[6].value
        if sheet_name == "konsensbasiert":
            continue
            assessment = assess_cb(reco_text, background_text, reco_gor)
            dict_idx = 9
        else:
            assessment = assess_eb(reco_text, background_text, reco_gor)
            dict_idx = 9
        for k,v in assessment.items():
            sheet.cell(row=row[0].row, column=dict_idx).value = int(v['grade'])
            dict_idx += 1
            sheet.cell(row=row[0].row, column=dict_idx).value = v['explanation']
            dict_idx += 1

workbook.save(target_path)